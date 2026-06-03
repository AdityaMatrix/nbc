from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone
from io import BytesIO
import uuid
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import db
from config import UserRole, PLANTS, DEPARTMENTS
from dependencies import get_current_user
from helpers import notify_users

router = APIRouter(prefix="/api")
logger = logging.getLogger(__name__)

# ---- Column definitions ----

# Columns for creating NEW requests
NEW_REQUEST_COLUMNS = [
    {"key": "plant", "header": "Plant *", "width": 15, "required": True,
     "help": f"Options: {', '.join(PLANTS)}"},
    {"key": "department", "header": "Department *", "width": 25, "required": True,
     "help": f"Options: {', '.join(DEPARTMENTS[:8])}..."},
    {"key": "requirement_description", "header": "Requirement Description *", "width": 40, "required": True,
     "help": "Brief description of the item/equipment needed"},
    {"key": "quantity", "header": "Quantity", "width": 12, "required": False,
     "help": "Number of items (default: 1)"},
    {"key": "requirement_type", "header": "Requirement Type *", "width": 20, "required": True,
     "help": "Options: New, Replacement, Upgrade"},
    {"key": "asset_category", "header": "Asset Category", "width": 20, "required": False,
     "help": "Category of asset (e.g., Machinery, Equipment, IT)"},
    {"key": "cea_required", "header": "CEA Required", "width": 15, "required": False,
     "help": "Yes / No (default: No)"},
    {"key": "cea_type", "header": "CEA Type", "width": 15, "required": False,
     "help": "new / existing (only if CEA Required = Yes)"},
    {"key": "existing_cea_number", "header": "Existing CEA Number", "width": 20, "required": False,
     "help": "Only if CEA Type = existing"},
    {"key": "pr_available", "header": "PR Available", "width": 15, "required": False,
     "help": "Yes / No (default: No)"},
    {"key": "pr_number", "header": "PR Number", "width": 18, "required": False,
     "help": "Only if PR Available = Yes"},
    {"key": "dap_required", "header": "DAP Required", "width": 15, "required": False,
     "help": "Yes / No (default: No)"},
    {"key": "justification", "header": "Justification", "width": 40, "required": False,
     "help": "Business justification for the request"},
]

# Columns for UPDATING existing requests
UPDATE_REQUEST_COLUMNS = [
    {"key": "request_id", "header": "Request ID *", "width": 20, "required": True,
     "help": "e.g., JAI-RB-001"},
    {"key": "cea_number", "header": "CEA Number", "width": 18, "required": False, "help": ""},
    {"key": "cea_status", "header": "CEA Status", "width": 18, "required": False,
     "help": "Capex Head, Department Head, CTO, Manufacturing Head, Operation Head, Budget, CFO, Approved"},
    {"key": "wbs_number", "header": "WBS Number", "width": 18, "required": False, "help": ""},
    {"key": "pr_number", "header": "PR Number", "width": 18, "required": False, "help": ""},
    {"key": "pr_approval_status", "header": "PR Status", "width": 18, "required": False,
     "help": "01-06, Approved"},
    {"key": "po_number", "header": "PO Number", "width": 18, "required": False, "help": ""},
    {"key": "po_approval_status", "header": "PO Status", "width": 18, "required": False,
     "help": "01-05, Approved"},
    {"key": "vendor_name", "header": "Vendor Name", "width": 25, "required": False, "help": ""},
    {"key": "initial_price", "header": "Initial Price", "width": 18, "required": False,
     "help": "Supplier quoted price (number)"},
    {"key": "final_price", "header": "Final Negotiated Price", "width": 20, "required": False,
     "help": "Negotiated price (number)"},
    {"key": "ordered_date", "header": "Ordered Date", "width": 18, "required": False,
     "help": "YYYY-MM-DD"},
    {"key": "expected_delivery_date", "header": "Expected Delivery Date", "width": 22, "required": False,
     "help": "YYYY-MM-DD"},
    {"key": "delivery_status", "header": "Delivery Status", "width": 20, "required": False,
     "help": "Yet to Dispatch, Dispatched, Delivered"},
    {"key": "delivery_date", "header": "Delivery Date", "width": 18, "required": False,
     "help": "YYYY-MM-DD"},
    {"key": "installation_date", "header": "Installation Date", "width": 18, "required": False,
     "help": "YYYY-MM-DD"},
    {"key": "commissioning_date", "header": "Commissioning Date", "width": 20, "required": False,
     "help": "YYYY-MM-DD"},
    {"key": "workflow_status", "header": "Workflow Status", "width": 25, "required": False,
     "help": "CEA Under Approval, PR Approved, Order Placed, Completed, etc."},
]


def _build_template_workbook():
    """Build the Excel template with two sheets: New Requests and Update Requests."""
    wb = Workbook()

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    help_font = Font(name="Calibri", italic=True, size=9, color="6B7280")
    help_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    required_fill = PatternFill(start_color="3730A3", end_color="3730A3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Sheet 1: New Requests
    ws_new = wb.active
    ws_new.title = "New Requests"
    ws_new.sheet_properties.tabColor = "4F46E5"

    for col_idx, col_def in enumerate(NEW_REQUEST_COLUMNS, 1):
        letter = get_column_letter(col_idx)
        # Header row
        cell = ws_new.cell(row=1, column=col_idx, value=col_def["header"])
        cell.font = header_font
        cell.fill = required_fill if col_def["required"] else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        # Help row
        help_cell = ws_new.cell(row=2, column=col_idx, value=col_def["help"])
        help_cell.font = help_font
        help_cell.fill = help_fill
        help_cell.alignment = Alignment(wrap_text=True, vertical="top")
        help_cell.border = thin_border
        # Column width
        ws_new.column_dimensions[letter].width = col_def["width"]

    ws_new.row_dimensions[1].height = 30
    ws_new.row_dimensions[2].height = 45
    ws_new.freeze_panes = "A3"

    # Add 2 sample rows
    samples = [
        ["Jaipur", "Railway Bearing", "CNC Milling Machine", 1, "New", "Machinery", "Yes", "new", "", "No", "", "No", "Required for production line upgrade"],
        ["Bagru", "Industrial Bearing", "Hydraulic Press 100T", 2, "Replacement", "Equipment", "No", "", "", "Yes", "PR-2026-001", "Yes", "Replacing old press for safety compliance"],
    ]
    for row_idx, sample in enumerate(samples, 3):
        for col_idx, val in enumerate(sample, 1):
            cell = ws_new.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top")

    # Sheet 2: Update Requests
    ws_update = wb.create_sheet("Update Existing Requests")
    ws_update.sheet_properties.tabColor = "059669"

    update_header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    update_req_fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")

    for col_idx, col_def in enumerate(UPDATE_REQUEST_COLUMNS, 1):
        letter = get_column_letter(col_idx)
        cell = ws_update.cell(row=1, column=col_idx, value=col_def["header"])
        cell.font = header_font
        cell.fill = update_req_fill if col_def["required"] else update_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        help_cell = ws_update.cell(row=2, column=col_idx, value=col_def["help"])
        help_cell.font = help_font
        help_cell.fill = help_fill
        help_cell.alignment = Alignment(wrap_text=True, vertical="top")
        help_cell.border = thin_border
        ws_update.column_dimensions[letter].width = col_def["width"]

    ws_update.row_dimensions[1].height = 30
    ws_update.row_dimensions[2].height = 45
    ws_update.freeze_panes = "A3"

    # Instructions sheet
    ws_info = wb.create_sheet("Instructions", 0)
    ws_info.sheet_properties.tabColor = "F59E0B"
    ws_info.column_dimensions["A"].width = 80

    instructions = [
        ("CAPEX PORTAL - BULK UPLOAD TEMPLATE", True, 16, "4F46E5"),
        ("", False, 11, None),
        ("HOW TO USE:", True, 13, "1E293B"),
        ("1. To CREATE new requests, fill in the 'New Requests' sheet", False, 11, None),
        ("2. To UPDATE existing requests, fill in the 'Update Existing Requests' sheet", False, 11, None),
        ("3. You can use both sheets in a single upload", False, 11, None),
        ("4. Row 2 contains help text/valid options - do NOT delete it", False, 11, None),
        ("5. Fields marked with * are required", False, 11, None),
        ("6. Leave cells empty if not applicable", False, 11, None),
        ("", False, 11, None),
        ("RULES:", True, 13, "1E293B"),
        ("- Plant must be one of: " + ", ".join(PLANTS), False, 11, None),
        ("- Department must be a valid department name", False, 11, None),
        ("- Dates must be in YYYY-MM-DD format", False, 11, None),
        ("- Prices must be numeric values (no currency symbols)", False, 11, None),
        ("- Yes/No fields accept: Yes, No, Y, N, TRUE, FALSE", False, 11, None),
        ("- For updates, only the Request ID is required; fill other columns only for fields you want to change", False, 11, None),
    ]

    for row_idx, (text, bold, size, color) in enumerate(instructions, 1):
        cell = ws_info.cell(row=row_idx, column=1, value=text)
        cell.font = Font(name="Calibri", bold=bold, size=size, color=color or "374151")
        cell.alignment = Alignment(wrap_text=True)

    return wb


def _parse_bool(val):
    if val is None or val == "":
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    return s in ("yes", "y", "true", "1")


def _clean_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _clean_number(val):
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _clean_date(val):
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # Try common formats
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


@router.get("/capex-requests/bulk-template")
async def download_bulk_template(current_user: dict = Depends(get_current_user)):
    """Download the Excel template for bulk upload."""
    wb = _build_template_workbook()
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Capex_Bulk_Upload_Template.xlsx"}
    )


@router.post("/capex-requests/bulk-upload")
async def bulk_upload_capex(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Process bulk upload of Capex requests (create new + update existing)."""
    from openpyxl import load_workbook

    if not file.filename.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, or .csv files are supported")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:  # 10MB limit
        raise HTTPException(status_code=400, detail="File size exceeds 10MB limit")

    results = {"created": [], "updated": [], "errors": [], "summary": {}}

    try:
        if file.filename.endswith(".csv"):
            import csv
            from io import StringIO
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(StringIO(text))
            rows = list(reader)
            # CSV: treat as new requests if no "Request ID" column, else updates
            for row_idx, row in enumerate(rows, 2):
                req_id = row.get("Request ID *") or row.get("Request ID") or row.get("request_id")
                if req_id and str(req_id).strip():
                    result = await _process_update_row(row_idx, row, current_user, is_csv=True)
                else:
                    result = await _process_new_row(row_idx, row, current_user, is_csv=True)
                if result.get("success"):
                    if result["action"] == "created":
                        results["created"].append(result)
                    else:
                        results["updated"].append(result)
                else:
                    results["errors"].append(result)
        else:
            wb = load_workbook(BytesIO(content), data_only=True)

            # Process "New Requests" sheet
            if "New Requests" in wb.sheetnames:
                ws = wb["New Requests"]
                headers = [cell.value for cell in ws[1]]
                for row_idx in range(3, ws.max_row + 1):  # Skip header + help rows
                    row_values = [cell.value for cell in ws[row_idx]]
                    if not any(v for v in row_values):
                        continue  # Skip empty rows
                    row_dict = dict(zip(headers, row_values))
                    result = await _process_new_row(row_idx, row_dict, current_user)
                    if result.get("success"):
                        results["created"].append(result)
                    else:
                        results["errors"].append(result)

            # Process "Update Existing Requests" sheet
            update_sheet_name = None
            for name in wb.sheetnames:
                if "update" in name.lower():
                    update_sheet_name = name
                    break
            if update_sheet_name:
                ws = wb[update_sheet_name]
                headers = [cell.value for cell in ws[1]]
                for row_idx in range(3, ws.max_row + 1):
                    row_values = [cell.value for cell in ws[row_idx]]
                    if not any(v for v in row_values):
                        continue
                    row_dict = dict(zip(headers, row_values))
                    result = await _process_update_row(row_idx, row_dict, current_user)
                    if result.get("success"):
                        results["updated"].append(result)
                    else:
                        results["errors"].append(result)

    except Exception as e:
        logger.error(f"Bulk upload parse error: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    results["summary"] = {
        "total_processed": len(results["created"]) + len(results["updated"]) + len(results["errors"]),
        "created": len(results["created"]),
        "updated": len(results["updated"]),
        "errors": len(results["errors"]),
    }

    # Save upload history log
    if results["created"] or results["updated"]:
        upload_log = {
            "id": str(uuid.uuid4()),
            "user_id": current_user["id"],
            "user_name": current_user["name"],
            "filename": file.filename,
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "created_request_ids": [r["request_id"] for r in results["created"]],
            "updated_snapshots": [r.get("_snapshot") for r in results["updated"] if r.get("_snapshot")],
            "summary": results["summary"],
            "status": "active",
        }
        await db.bulk_upload_logs.insert_one(upload_log)
        results["upload_id"] = upload_log["id"]

    # Strip internal snapshot data from response
    for item in results["updated"]:
        item.pop("_snapshot", None)

    return results


async def _process_new_row(row_idx, row_dict, current_user, is_csv=False):
    """Create a new capex request from a row."""
    try:
        # Extract fields
        plant = _clean_str(row_dict.get("Plant *") or row_dict.get("plant"))
        department = _clean_str(row_dict.get("Department *") or row_dict.get("department"))
        description = _clean_str(row_dict.get("Requirement Description *") or row_dict.get("requirement_description"))
        quantity = row_dict.get("Quantity") or row_dict.get("quantity") or 1
        req_type = _clean_str(row_dict.get("Requirement Type *") or row_dict.get("requirement_type"))
        asset_category = _clean_str(row_dict.get("Asset Category") or row_dict.get("asset_category"))
        cea_required = _parse_bool(row_dict.get("CEA Required") or row_dict.get("cea_required"))
        cea_type = _clean_str(row_dict.get("CEA Type") or row_dict.get("cea_type"))
        existing_cea = _clean_str(row_dict.get("Existing CEA Number") or row_dict.get("existing_cea_number"))
        pr_available = _parse_bool(row_dict.get("PR Available") or row_dict.get("pr_available"))
        pr_number = _clean_str(row_dict.get("PR Number") or row_dict.get("pr_number"))
        dap_required = _parse_bool(row_dict.get("DAP Required") or row_dict.get("dap_required"))
        justification = _clean_str(row_dict.get("Justification") or row_dict.get("justification"))

        # Validate required
        errors = []
        if not plant:
            errors.append("Plant is required")
        elif plant not in PLANTS:
            errors.append(f"Invalid plant: {plant}")
        if not department:
            errors.append("Department is required")
        elif department not in DEPARTMENTS:
            errors.append(f"Invalid department: {department}")
        if not description:
            errors.append("Requirement Description is required")
        if not req_type:
            errors.append("Requirement Type is required")

        if errors:
            return {"success": False, "row": row_idx, "action": "create", "errors": errors}

        # Build request
        try:
            quantity = int(quantity)
        except (ValueError, TypeError):
            quantity = 1

        plant_code = plant[:3].upper()
        dept_words = department.split()
        dept_code = "".join([w[0].upper() for w in dept_words if w])
        total_count = await db.capex_requests.count_documents({})
        serial_num = str(total_count + 1).zfill(3)
        request_id = f"{plant_code}-{dept_code}-{serial_num}"

        initial_workflow_status = None
        initial_cea_status = None
        wbs_number = None

        if cea_required:
            if cea_type == "new":
                initial_workflow_status = "CEA Under Approval"
            elif cea_type == "existing" and existing_cea:
                wbs_number = existing_cea
                initial_cea_status = "Approved"
                initial_workflow_status = "CEA Approved"

        initial_pr_approval_status = None
        if pr_available and pr_number:
            initial_pr_approval_status = "Approved"
            initial_workflow_status = "PR Approved"

        capex_doc = {
            "id": request_id,
            "user_id": current_user["id"],
            "user_name": current_user["name"],
            "user_email": current_user["email"],
            "plant": plant,
            "department": department,
            "asset_category": asset_category,
            "requirement_items": [{"description": description, "quantity": quantity}],
            "requirement_description": description,
            "requirement_type": req_type,
            "cea_required": cea_required,
            "cea_type": cea_type if cea_required else None,
            "cea_number": None,
            "wbs_number": wbs_number,
            "cea_status": initial_cea_status,
            "pr_available": pr_available,
            "pr_number": pr_number if pr_available else None,
            "pr_approval_status": initial_pr_approval_status,
            "dap_required": dap_required,
            "justification": justification,
            "status": "Pending DH Approval",
            "workflow_status": initial_workflow_status,
            "current_workflow_stage": "request",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "time_log": [{
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "action": "Request created via bulk upload",
                "by": current_user["name"],
                "by_id": current_user["id"]
            }],
            "attachments": [],
            "cea_creation_date": None,
            "pr_processed": False, "pr_approval_level": None,
            "po_available": False, "po_processed": False, "po_number": None,
            "po_approval_status": None, "po_approval_level": None,
            "expected_delivery_date": None, "installation_date": None,
            "commissioning_date": None, "commissioning_status": None,
            "assigned_buyer_id": None, "assigned_buyer_name": None,
            "dap_id": None,
            "dh_approval_status": "Pending",
            "dh_approved_by": None, "dh_approved_at": None, "dh_rejection_reason": None,
            "pdi_status": None, "pdi_date": None, "pdi_remarks": None,
            "delivery_status": None, "delivery_date": None,
            "suppliers": [],
            "payment_terms": [],
            "gst_applicable": None,
            "gst_percentage": None,
            "pr_provided_by": "user" if pr_available and pr_number else None,
        }

        await db.capex_requests.insert_one(capex_doc)

        return {
            "success": True,
            "row": row_idx,
            "action": "created",
            "request_id": request_id,
            "description": description,
        }

    except Exception as e:
        logger.error(f"Row {row_idx} create error: {e}")
        return {"success": False, "row": row_idx, "action": "create", "errors": [str(e)]}


async def _process_update_row(row_idx, row_dict, current_user, is_csv=False):
    """Update an existing capex request from a row."""
    try:
        request_id = _clean_str(row_dict.get("Request ID *") or row_dict.get("Request ID") or row_dict.get("request_id"))
        if not request_id:
            return {"success": False, "row": row_idx, "action": "update", "errors": ["Request ID is required"]}

        existing = await db.capex_requests.find_one({"id": request_id}, {"_id": 0})
        if not existing:
            return {"success": False, "row": row_idx, "action": "update", "errors": [f"Request '{request_id}' not found"]}

        update_data = {}

        # Map columns to fields
        field_map = {
            "CEA Number": "cea_number",
            "CEA Status": "cea_status",
            "WBS Number": "wbs_number",
            "PR Number": "pr_number",
            "PR Status": "pr_approval_status",
            "PO Number": "po_number",
            "PO Status": "po_approval_status",
            "Vendor Name": "vendor_name",
            "Delivery Status": "delivery_status",
            "Workflow Status": "workflow_status",
        }

        for header, field in field_map.items():
            val = _clean_str(row_dict.get(header) or row_dict.get(field))
            if val:
                update_data[field] = val

        # Numeric fields
        for header, field in [("Initial Price", "initial_price"), ("Final Negotiated Price", "final_price")]:
            val = _clean_number(row_dict.get(header) or row_dict.get(field))
            if val is not None:
                update_data[field] = val

        # Date fields
        date_map = {
            "Ordered Date": "ordered_date",
            "Expected Delivery Date": "expected_delivery_date",
            "Delivery Date": "delivery_date",
            "Installation Date": "installation_date",
            "Commissioning Date": "commissioning_date",
        }
        for header, field in date_map.items():
            val = _clean_date(row_dict.get(header) or row_dict.get(field))
            if val:
                update_data[field] = val

        if not update_data:
            return {"success": False, "row": row_idx, "action": "update", "errors": ["No fields to update"]}

        # Handle supplier price updates
        if "initial_price" in update_data or "final_price" in update_data:
            suppliers = existing.get("suppliers", [])
            vendor_name = update_data.pop("vendor_name", None) or existing.get("vendor_name")
            initial = update_data.pop("initial_price", None)
            final = update_data.pop("final_price", None)

            if suppliers:
                # Update the first ordered/selected supplier
                target = next(
                    (s for s in suppliers if s.get("is_ordered")),
                    next((s for s in suppliers if s.get("selected")), suppliers[0] if suppliers else None)
                )
                if target:
                    if initial is not None:
                        target["initial_price"] = initial
                    if final is not None:
                        target["final_price"] = final
                    if vendor_name:
                        target["name"] = vendor_name
                    update_data["suppliers"] = suppliers
            else:
                # Create a new supplier entry
                new_supplier = {
                    "name": vendor_name or "Vendor (Bulk Upload)",
                    "initial_price": initial or 0,
                    "final_price": final or 0,
                    "selected": True,
                    "is_ordered": True if update_data.get("ordered_date") else False,
                }
                update_data["suppliers"] = [new_supplier]
        elif "vendor_name" in update_data:
            update_data.pop("vendor_name")

        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()

        time_log_entry = {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "action": f"Bulk update: {', '.join(k for k in update_data.keys() if k != 'updated_at')}",
            "by": current_user["name"],
            "by_id": current_user["id"]
        }

        await db.capex_requests.update_one(
            {"id": request_id},
            {"$set": update_data, "$push": {"time_log": time_log_entry}}
        )

        # Build snapshot of previous values for rollback
        snapshot_fields = [k for k in update_data.keys() if k not in ("updated_at",)]
        previous_values = {}
        for field in snapshot_fields:
            if field in existing:
                previous_values[field] = existing[field]
            else:
                previous_values[field] = None

        return {
            "success": True,
            "row": row_idx,
            "action": "updated",
            "request_id": request_id,
            "fields_updated": snapshot_fields,
            "_snapshot": {
                "request_id": request_id,
                "previous_values": previous_values,
            },
        }

    except Exception as e:
        logger.error(f"Row {row_idx} update error: {e}")
        return {"success": False, "row": row_idx, "action": "update", "errors": [str(e)]}


@router.get("/capex-requests/bulk-upload/history")
async def get_bulk_upload_history(current_user: dict = Depends(get_current_user)):
    """Get bulk upload history for the current user (Capex Head sees all)."""
    query = {}
    if current_user["role"] != UserRole.CAPEX_HEAD:
        query["user_id"] = current_user["id"]

    logs = await db.bulk_upload_logs.find(query, {"_id": 0}).sort("uploaded_at", -1).limit(50).to_list(50)
    return logs


@router.post("/capex-requests/bulk-upload/{upload_id}/rollback")
async def rollback_bulk_upload(upload_id: str, current_user: dict = Depends(get_current_user)):
    """Rollback a bulk upload: delete created requests, revert updated requests."""
    log_entry = await db.bulk_upload_logs.find_one({"id": upload_id}, {"_id": 0})
    if not log_entry:
        raise HTTPException(status_code=404, detail="Upload log not found")

    if log_entry.get("status") == "rolled_back":
        raise HTTPException(status_code=400, detail="This upload has already been rolled back")

    # Permission check: only the uploader or Capex Head can rollback
    if log_entry["user_id"] != current_user["id"] and current_user["role"] != UserRole.CAPEX_HEAD:
        raise HTTPException(status_code=403, detail="Not authorized to rollback this upload")

    rollback_results = {"deleted": [], "reverted": [], "errors": []}

    # 1. Delete created requests
    for request_id in log_entry.get("created_request_ids", []):
        try:
            existing = await db.capex_requests.find_one({"id": request_id}, {"_id": 0})
            if existing:
                await db.capex_requests.delete_one({"id": request_id})
                await db.comments.delete_many({"capex_request_id": request_id})
                await db.sample_requests.delete_many({"capex_request_id": request_id})
                rollback_results["deleted"].append(request_id)
            else:
                rollback_results["errors"].append(f"{request_id}: already deleted")
        except Exception as e:
            rollback_results["errors"].append(f"{request_id}: {str(e)}")

    # 2. Revert updated requests
    for snapshot in log_entry.get("updated_snapshots", []):
        request_id = snapshot["request_id"]
        previous_values = snapshot["previous_values"]
        try:
            existing = await db.capex_requests.find_one({"id": request_id}, {"_id": 0})
            if not existing:
                rollback_results["errors"].append(f"{request_id}: not found")
                continue

            revert_data = {}
            unset_data = {}
            for field, prev_val in previous_values.items():
                if prev_val is None:
                    unset_data[field] = ""
                else:
                    revert_data[field] = prev_val

            revert_data["updated_at"] = datetime.now(timezone.utc).isoformat()

            update_ops = {"$set": revert_data}
            if unset_data:
                update_ops["$unset"] = unset_data
            update_ops.setdefault("$push", {})["time_log"] = {
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "action": f"Bulk upload rollback: reverted {', '.join(previous_values.keys())}",
                "by": current_user["name"],
                "by_id": current_user["id"]
            }

            await db.capex_requests.update_one({"id": request_id}, update_ops)
            rollback_results["reverted"].append(request_id)
        except Exception as e:
            rollback_results["errors"].append(f"{request_id}: {str(e)}")

    # Mark the log as rolled back
    await db.bulk_upload_logs.update_one({"id": upload_id}, {
        "$set": {
            "status": "rolled_back",
            "rolled_back_at": datetime.now(timezone.utc).isoformat(),
            "rolled_back_by": current_user["name"],
        }
    })

    return {
        "message": "Rollback completed",
        "upload_id": upload_id,
        "deleted": len(rollback_results["deleted"]),
        "reverted": len(rollback_results["reverted"]),
        "errors": rollback_results["errors"],
    }
