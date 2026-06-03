"""
Test suite for Bulk Upload History and Rollback features - Capex Portal
Tests: Upload history log, rollback created requests, rollback updated requests, double rollback prevention
"""
import pytest
import requests
import os
import io
import time
from openpyxl import Workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Test credentials
CAPEX_HEAD_EMAIL = "manoj@capex.com"
CAPEX_HEAD_PASSWORD = "capex123"
BUYER_EMAIL = "vijay@capex.com"
BUYER_PASSWORD = "buyer123"

# Valid values from config
VALID_PLANTS = ["Bagru", "Jaipur", "Newai", "Savli"]
VALID_DEPARTMENTS = [
    "Railway Bearing", "Industrial Bearing", "Ball Bearing", "Taper Roller Bearing",
    "Large Dia Bearing", "Water Pump Bearing", "Finish Goods", "Stores",
    "Digital", "Quality", "IT", "HR", "Marketing & Branding", "R&D", "Metallurgy Lab", "Tribo Lab"
]


@pytest.fixture(scope="module")
def api_client():
    """Shared requests session"""
    session = requests.Session()
    return session


@pytest.fixture(scope="module")
def capex_head_token(api_client):
    """Get authentication token for Capex Head"""
    response = api_client.post(f"{BASE_URL}/api/auth/login", 
        headers={"Content-Type": "application/json"},
        json={
            "email": CAPEX_HEAD_EMAIL,
            "password": CAPEX_HEAD_PASSWORD
        })
    if response.status_code == 200:
        return response.json().get("access_token")
    pytest.skip(f"Capex Head authentication failed: {response.status_code} - {response.text}")


@pytest.fixture(scope="module")
def buyer_token(api_client):
    """Get authentication token for Buyer"""
    response = api_client.post(f"{BASE_URL}/api/auth/login", 
        headers={"Content-Type": "application/json"},
        json={
            "email": BUYER_EMAIL,
            "password": BUYER_PASSWORD
        })
    if response.status_code == 200:
        return response.json().get("access_token")
    pytest.skip(f"Buyer authentication failed: {response.status_code}")


# Track created upload IDs for cleanup
created_upload_ids = []
created_request_ids = []


@pytest.fixture(scope="module", autouse=True)
def cleanup_test_data(api_client, capex_head_token):
    """Cleanup test-created data after all tests complete"""
    yield
    # Teardown: Delete all test-created requests and upload logs
    if capex_head_token:
        headers = {"Authorization": f"Bearer {capex_head_token}"}
        for req_id in created_request_ids:
            try:
                api_client.delete(f"{BASE_URL}/api/capex-requests/{req_id}", headers=headers)
                print(f"Cleaned up test request: {req_id}")
            except Exception as e:
                print(f"Failed to cleanup request {req_id}: {e}")


def create_test_excel_for_new_requests(descriptions):
    """Helper to create Excel file with new requests"""
    wb = Workbook()
    ws = wb.active
    ws.title = "New Requests"
    
    headers = ["Plant *", "Department *", "Requirement Description *", "Quantity", 
               "Requirement Type *", "Asset Category", "CEA Required", "CEA Type",
               "Existing CEA Number", "PR Available", "PR Number", "DAP Required", "Justification"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    for col in range(1, len(headers) + 1):
        ws.cell(row=2, column=col, value="")
    
    for row_idx, desc in enumerate(descriptions, 3):
        data = ["Jaipur", "Railway Bearing", desc, 1, "New", "Machinery", "No", "", "", "No", "", "No", "Test"]
        for col, val in enumerate(data, 1):
            ws.cell(row=row_idx, column=col, value=val)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_test_excel_for_updates(updates):
    """Helper to create Excel file with update requests"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Update Existing Requests"
    
    update_headers = ["Request ID *", "CEA Number", "CEA Status", "WBS Number", "PR Number", 
                      "PR Status", "PO Number", "PO Status", "Vendor Name", "Initial Price",
                      "Final Negotiated Price", "Ordered Date", "Expected Delivery Date",
                      "Delivery Status", "Delivery Date", "Installation Date", "Commissioning Date", "Workflow Status"]
    for col, header in enumerate(update_headers, 1):
        ws.cell(row=1, column=col, value=header)
    for col in range(1, len(update_headers) + 1):
        ws.cell(row=2, column=col, value="")
    
    for row_idx, update in enumerate(updates, 3):
        for col, val in enumerate(update, 1):
            ws.cell(row=row_idx, column=col, value=val)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


class TestBulkUploadHistoryEndpoint:
    """Tests for GET /api/capex-requests/bulk-upload/history"""
    
    def test_history_endpoint_returns_list(self, api_client, capex_head_token):
        """History endpoint returns a list of upload logs"""
        response = api_client.get(
            f"{BASE_URL}/api/capex-requests/bulk-upload/history",
            headers={"Authorization": f"Bearer {capex_head_token}"}
        )
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        data = response.json()
        assert isinstance(data, list), "Response should be a list"
        print(f"History endpoint returned {len(data)} logs")
    
    def test_history_requires_authentication(self, api_client):
        """History endpoint requires authentication"""
        response = api_client.get(f"{BASE_URL}/api/capex-requests/bulk-upload/history")
        
        assert response.status_code in [401, 403], \
            f"Expected 401/403 for unauthorized, got {response.status_code}"
    
    def test_upload_creates_history_log(self, api_client, capex_head_token):
        """Bulk upload creates a history log entry"""
        # Create and upload a file
        buffer = create_test_excel_for_new_requests(["TEST_History_Log_Machine_001"])
        
        upload_response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload",
            headers={"Authorization": f"Bearer {capex_head_token}"},
            files={"file": ("test_history.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert upload_response.status_code == 200, f"Upload failed: {upload_response.text}"
        upload_result = upload_response.json()
        
        # Track for cleanup
        for created in upload_result.get("created", []):
            if created.get("request_id"):
                created_request_ids.append(created["request_id"])
        
        # Check upload_id is returned
        upload_id = upload_result.get("upload_id")
        assert upload_id is not None, "Upload should return upload_id"
        created_upload_ids.append(upload_id)
        
        # Verify history contains this upload
        history_response = api_client.get(
            f"{BASE_URL}/api/capex-requests/bulk-upload/history",
            headers={"Authorization": f"Bearer {capex_head_token}"}
        )
        
        assert history_response.status_code == 200
        history = history_response.json()
        
        # Find our upload in history
        our_log = next((log for log in history if log.get("id") == upload_id), None)
        assert our_log is not None, f"Upload {upload_id} should be in history"
        
        # Verify log metadata
        assert our_log.get("filename") == "test_history.xlsx", "Filename should be stored"
        assert our_log.get("user_name") is not None, "User name should be stored"
        assert our_log.get("uploaded_at") is not None, "Upload timestamp should be stored"
        assert our_log.get("status") == "active", "Status should be 'active'"
        assert our_log.get("summary") is not None, "Summary should be stored"
        assert len(our_log.get("created_request_ids", [])) >= 1, "Created request IDs should be stored"
        
        print(f"History log created successfully: {upload_id}")


class TestRollbackCreatedRequests:
    """Tests for POST /api/capex-requests/bulk-upload/{upload_id}/rollback - deleting created requests"""
    
    def test_rollback_deletes_created_requests(self, api_client, capex_head_token):
        """Rollback deletes requests that were created in the upload"""
        # Create requests via bulk upload
        buffer = create_test_excel_for_new_requests([
            "TEST_Rollback_Delete_001",
            "TEST_Rollback_Delete_002"
        ])
        
        upload_response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload",
            headers={"Authorization": f"Bearer {capex_head_token}"},
            files={"file": ("test_rollback_create.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert upload_response.status_code == 200
        upload_result = upload_response.json()
        upload_id = upload_result.get("upload_id")
        assert upload_id is not None, "Upload should return upload_id"
        
        created_ids = [c["request_id"] for c in upload_result.get("created", [])]
        assert len(created_ids) == 2, f"Should create 2 requests, got {len(created_ids)}"
        
        # Verify requests exist
        for req_id in created_ids:
            get_response = api_client.get(
                f"{BASE_URL}/api/capex-requests/{req_id}",
                headers={"Authorization": f"Bearer {capex_head_token}"}
            )
            assert get_response.status_code == 200, f"Request {req_id} should exist before rollback"
        
        # Perform rollback
        rollback_response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload/{upload_id}/rollback",
            headers={"Authorization": f"Bearer {capex_head_token}"}
        )
        
        assert rollback_response.status_code == 200, f"Rollback failed: {rollback_response.text}"
        rollback_result = rollback_response.json()
        
        assert rollback_result.get("deleted") == 2, f"Should delete 2 requests, got {rollback_result.get('deleted')}"
        assert rollback_result.get("upload_id") == upload_id, "Should return upload_id"
        
        # Verify requests are deleted
        for req_id in created_ids:
            get_response = api_client.get(
                f"{BASE_URL}/api/capex-requests/{req_id}",
                headers={"Authorization": f"Bearer {capex_head_token}"}
            )
            assert get_response.status_code == 404, f"Request {req_id} should be deleted after rollback"
        
        print(f"Rollback successfully deleted {len(created_ids)} requests")


class TestRollbackUpdatedRequests:
    """Tests for rollback reverting updated requests to previous values"""
    
    def test_rollback_reverts_updated_requests(self, api_client, capex_head_token):
        """Rollback reverts updated requests to their previous values"""
        # First, create a request to update
        buffer_create = create_test_excel_for_new_requests(["TEST_Rollback_Revert_Target"])
        
        create_response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload",
            headers={"Authorization": f"Bearer {capex_head_token}"},
            files={"file": ("create_for_revert.xlsx", buffer_create, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert create_response.status_code == 200
        create_result = create_response.json()
        request_id = create_result["created"][0]["request_id"]
        created_request_ids.append(request_id)
        
        # Get original values
        original_response = api_client.get(
            f"{BASE_URL}/api/capex-requests/{request_id}",
            headers={"Authorization": f"Bearer {capex_head_token}"}
        )
        original_data = original_response.json()
        original_pr_number = original_data.get("pr_number")
        original_workflow_status = original_data.get("workflow_status")
        
        # Update the request via bulk upload
        update_data = [request_id, "", "", "", "PR-ROLLBACK-TEST", "Approved", "", "", "", "", "", "", "", "", "", "", "", "PR Approved"]
        buffer_update = create_test_excel_for_updates([update_data])
        
        update_response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload",
            headers={"Authorization": f"Bearer {capex_head_token}"},
            files={"file": ("update_for_revert.xlsx", buffer_update, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert update_response.status_code == 200
        update_result = update_response.json()
        upload_id = update_result.get("upload_id")
        assert upload_id is not None, "Update should return upload_id"
        assert update_result["summary"]["updated"] >= 1, "Should update at least 1 request"
        
        # Verify update was applied
        updated_response = api_client.get(
            f"{BASE_URL}/api/capex-requests/{request_id}",
            headers={"Authorization": f"Bearer {capex_head_token}"}
        )
        updated_data = updated_response.json()
        assert updated_data.get("pr_number") == "PR-ROLLBACK-TEST", "PR number should be updated"
        assert updated_data.get("workflow_status") == "PR Approved", "Workflow status should be updated"
        
        # Perform rollback
        rollback_response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload/{upload_id}/rollback",
            headers={"Authorization": f"Bearer {capex_head_token}"}
        )
        
        assert rollback_response.status_code == 200, f"Rollback failed: {rollback_response.text}"
        rollback_result = rollback_response.json()
        assert rollback_result.get("reverted") >= 1, f"Should revert at least 1 request, got {rollback_result.get('reverted')}"
        
        # Verify values are reverted
        reverted_response = api_client.get(
            f"{BASE_URL}/api/capex-requests/{request_id}",
            headers={"Authorization": f"Bearer {capex_head_token}"}
        )
        reverted_data = reverted_response.json()
        assert reverted_data.get("pr_number") == original_pr_number, \
            f"PR number should be reverted to {original_pr_number}, got {reverted_data.get('pr_number')}"
        assert reverted_data.get("workflow_status") == original_workflow_status, \
            f"Workflow status should be reverted to {original_workflow_status}, got {reverted_data.get('workflow_status')}"
        
        print(f"Rollback successfully reverted request {request_id} to original values")


class TestRollbackStatusAndMetadata:
    """Tests for rollback status updates and metadata"""
    
    def test_rollback_marks_log_as_rolled_back(self, api_client, capex_head_token):
        """Rollback marks the upload log as 'rolled_back' with timestamp and user"""
        # Create a request
        buffer = create_test_excel_for_new_requests(["TEST_Rollback_Status_Check"])
        
        upload_response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload",
            headers={"Authorization": f"Bearer {capex_head_token}"},
            files={"file": ("test_status.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert upload_response.status_code == 200
        upload_id = upload_response.json().get("upload_id")
        
        # Perform rollback
        rollback_response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload/{upload_id}/rollback",
            headers={"Authorization": f"Bearer {capex_head_token}"}
        )
        
        assert rollback_response.status_code == 200
        
        # Check history for updated status
        history_response = api_client.get(
            f"{BASE_URL}/api/capex-requests/bulk-upload/history",
            headers={"Authorization": f"Bearer {capex_head_token}"}
        )
        
        history = history_response.json()
        our_log = next((log for log in history if log.get("id") == upload_id), None)
        
        assert our_log is not None, "Log should still exist in history"
        assert our_log.get("status") == "rolled_back", f"Status should be 'rolled_back', got {our_log.get('status')}"
        assert our_log.get("rolled_back_at") is not None, "rolled_back_at timestamp should be set"
        assert our_log.get("rolled_back_by") is not None, "rolled_back_by should be set"
        
        print(f"Rollback status correctly updated: {our_log.get('status')}, by {our_log.get('rolled_back_by')}")


class TestDoubleRollbackPrevention:
    """Tests for preventing double rollback"""
    
    def test_double_rollback_returns_400(self, api_client, capex_head_token):
        """Attempting to rollback an already rolled back upload returns 400"""
        # Create a request
        buffer = create_test_excel_for_new_requests(["TEST_Double_Rollback_Check"])
        
        upload_response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload",
            headers={"Authorization": f"Bearer {capex_head_token}"},
            files={"file": ("test_double.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert upload_response.status_code == 200
        upload_id = upload_response.json().get("upload_id")
        
        # First rollback - should succeed
        first_rollback = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload/{upload_id}/rollback",
            headers={"Authorization": f"Bearer {capex_head_token}"}
        )
        assert first_rollback.status_code == 200, "First rollback should succeed"
        
        # Second rollback - should fail with 400
        second_rollback = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload/{upload_id}/rollback",
            headers={"Authorization": f"Bearer {capex_head_token}"}
        )
        
        assert second_rollback.status_code == 400, \
            f"Second rollback should return 400, got {second_rollback.status_code}"
        
        error_detail = second_rollback.json().get("detail", "")
        assert "already" in error_detail.lower() or "rolled back" in error_detail.lower(), \
            f"Error should mention already rolled back: {error_detail}"
        
        print(f"Double rollback correctly prevented with error: {error_detail}")


class TestRollbackNotFound:
    """Tests for rollback with non-existent upload ID"""
    
    def test_rollback_nonexistent_upload_returns_404(self, api_client, capex_head_token):
        """Rollback with non-existent upload ID returns 404"""
        response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload/nonexistent-upload-id-12345/rollback",
            headers={"Authorization": f"Bearer {capex_head_token}"}
        )
        
        assert response.status_code == 404, \
            f"Expected 404 for non-existent upload, got {response.status_code}"
        
        print("Non-existent upload rollback correctly returns 404")


class TestRollbackAuthorization:
    """Tests for rollback authorization"""
    
    def test_rollback_requires_authentication(self, api_client):
        """Rollback requires authentication"""
        response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload/some-upload-id/rollback"
        )
        
        assert response.status_code in [401, 403], \
            f"Expected 401/403 for unauthorized, got {response.status_code}"


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
