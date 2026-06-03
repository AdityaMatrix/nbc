"""
Test suite for Bulk Upload feature - Capex Portal
Tests: Template download, bulk create, bulk update, validation errors
"""
import pytest
import requests
import os
import io
from openpyxl import Workbook, load_workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Test credentials
CAPEX_HEAD_EMAIL = "manoj@capex.com"
CAPEX_HEAD_PASSWORD = "capex123"
USER_EMAIL = "amit@capex.com"
USER_PASSWORD = "user123"

# Valid values from config
VALID_PLANTS = ["Bagru", "Jaipur", "Newai", "Savli"]
VALID_DEPARTMENTS = [
    "Railway Bearing", "Industrial Bearing", "Ball Bearing", "Taper Roller Bearing",
    "Large Dia Bearing", "Water Pump Bearing", "Finish Goods", "Stores",
    "Digital", "Quality", "IT", "HR", "Marketing & Branding", "R&D", "Metallurgy Lab", "Tribo Lab"
]


@pytest.fixture(scope="module")
def api_client():
    """Shared requests session"""
    session = requests.Session()
    return session


@pytest.fixture(scope="module")
def auth_token(api_client):
    """Get authentication token for Capex Head"""
    response = api_client.post(f"{BASE_URL}/api/auth/login", 
        headers={"Content-Type": "application/json"},
        json={
            "email": CAPEX_HEAD_EMAIL,
            "password": CAPEX_HEAD_PASSWORD
        })
    if response.status_code == 200:
        return response.json().get("access_token")  # Note: access_token not token
    pytest.skip(f"Authentication failed: {response.status_code} - {response.text}")


@pytest.fixture(scope="module")
def user_token(api_client):
    """Get authentication token for regular user"""
    response = api_client.post(f"{BASE_URL}/api/auth/login", 
        headers={"Content-Type": "application/json"},
        json={
            "email": USER_EMAIL,
            "password": USER_PASSWORD
        })
    if response.status_code == 200:
        return response.json().get("access_token")  # Note: access_token not token
    pytest.skip(f"User authentication failed: {response.status_code}")


@pytest.fixture
def authenticated_client(api_client, auth_token):
    """Session with auth header"""
    api_client.headers.update({"Authorization": f"Bearer {auth_token}"})
    return api_client


@pytest.fixture
def user_authenticated_client(api_client, user_token):
    """Session with user auth header"""
    api_client.headers.update({"Authorization": f"Bearer {user_token}"})
    return api_client


# Track created request IDs for cleanup
created_request_ids = []


@pytest.fixture(scope="module", autouse=True)
def cleanup_test_data(api_client, auth_token):
    """Cleanup test-created data after all tests complete"""
    yield
    # Teardown: Delete all test-created requests
    if auth_token and created_request_ids:
        api_client.headers.update({"Authorization": f"Bearer {auth_token}"})
        for req_id in created_request_ids:
            try:
                api_client.delete(f"{BASE_URL}/api/capex-requests/{req_id}")
                print(f"Cleaned up test request: {req_id}")
            except Exception as e:
                print(f"Failed to cleanup {req_id}: {e}")


class TestBulkTemplateDownload:
    """Tests for GET /api/capex-requests/bulk-template"""
    
    def test_download_template_success(self, api_client, auth_token):
        """Template download returns Excel file with correct content-type"""
        response = api_client.get(
            f"{BASE_URL}/api/capex-requests/bulk-template",
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        
        # Check content type is Excel
        content_type = response.headers.get("Content-Type", "")
        assert "spreadsheetml" in content_type or "excel" in content_type, \
            f"Expected Excel content-type, got: {content_type}"
        
        # Check content-disposition header
        content_disp = response.headers.get("Content-Disposition", "")
        assert "attachment" in content_disp, "Expected attachment disposition"
        assert "Capex_Bulk_Upload_Template.xlsx" in content_disp, "Expected correct filename"
        
        # Verify it's a valid Excel file by loading it
        wb = load_workbook(io.BytesIO(response.content))
        sheet_names = wb.sheetnames
        
        # Should have 3 sheets: Instructions, New Requests, Update Existing Requests
        assert "Instructions" in sheet_names, "Missing Instructions sheet"
        assert "New Requests" in sheet_names, "Missing New Requests sheet"
        assert any("Update" in name for name in sheet_names), "Missing Update sheet"
        
        print(f"Template downloaded successfully with sheets: {sheet_names}")
    
    def test_download_template_unauthorized(self, api_client):
        """Template download requires authentication"""
        response = api_client.get(f"{BASE_URL}/api/capex-requests/bulk-template")
        
        assert response.status_code in [401, 403], \
            f"Expected 401/403 for unauthorized, got {response.status_code}"


class TestBulkUploadCreate:
    """Tests for POST /api/capex-requests/bulk-upload - Creating new requests"""
    
    def test_bulk_create_single_request(self, api_client, auth_token):
        """Upload Excel with single new request creates it successfully"""
        # Create Excel file with one valid row
        wb = Workbook()
        ws = wb.active
        ws.title = "New Requests"
        
        # Headers (row 1)
        headers = ["Plant *", "Department *", "Requirement Description *", "Quantity", 
                   "Requirement Type *", "Asset Category", "CEA Required", "CEA Type",
                   "Existing CEA Number", "PR Available", "PR Number", "DAP Required", "Justification"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Help row (row 2) - can be empty
        for col in range(1, len(headers) + 1):
            ws.cell(row=2, column=col, value="")
        
        # Data row (row 3)
        data = ["Jaipur", "Railway Bearing", "TEST_Bulk_Upload_Machine_001", 1, 
                "New", "Machinery", "No", "", "", "No", "", "No", "Test bulk upload"]
        for col, val in enumerate(data, 1):
            ws.cell(row=3, column=col, value=val)
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Upload
        response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload",
            headers={"Authorization": f"Bearer {auth_token}"},
            files={"file": ("test_upload.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        result = response.json()
        assert "summary" in result, "Response should have summary"
        assert result["summary"]["created"] >= 1, "Should have created at least 1 request"
        assert result["summary"]["errors"] == 0, f"Should have no errors: {result.get('errors', [])}"
        
        # Track for cleanup
        for created in result.get("created", []):
            if created.get("request_id"):
                created_request_ids.append(created["request_id"])
                print(f"Created request: {created['request_id']}")
    
    def test_bulk_create_multiple_requests(self, api_client, auth_token):
        """Upload Excel with multiple new requests creates them all"""
        wb = Workbook()
        ws = wb.active
        ws.title = "New Requests"
        
        headers = ["Plant *", "Department *", "Requirement Description *", "Quantity", 
                   "Requirement Type *", "Asset Category", "CEA Required", "CEA Type",
                   "Existing CEA Number", "PR Available", "PR Number", "DAP Required", "Justification"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Help row
        for col in range(1, len(headers) + 1):
            ws.cell(row=2, column=col, value="")
        
        # Multiple data rows
        test_data = [
            ["Bagru", "Industrial Bearing", "TEST_Bulk_Machine_A", 2, "Replacement", "Equipment", "No", "", "", "No", "", "No", "Test A"],
            ["Newai", "Ball Bearing", "TEST_Bulk_Machine_B", 1, "Upgrade", "Machinery", "Yes", "new", "", "No", "", "Yes", "Test B"],
            ["Savli", "IT", "TEST_Bulk_Computer", 5, "New", "IT Equipment", "No", "", "", "No", "", "No", "Test C"],
        ]
        
        for row_idx, data in enumerate(test_data, 3):
            for col, val in enumerate(data, 1):
                ws.cell(row=row_idx, column=col, value=val)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload",
            headers={"Authorization": f"Bearer {auth_token}"},
            files={"file": ("test_multi.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        
        result = response.json()
        assert result["summary"]["created"] == 3, f"Should create 3 requests, got {result['summary']['created']}"
        assert result["summary"]["errors"] == 0, f"Should have no errors: {result.get('errors', [])}"
        
        # Track for cleanup
        for created in result.get("created", []):
            if created.get("request_id"):
                created_request_ids.append(created["request_id"])
                print(f"Created request: {created['request_id']}")


class TestBulkUploadUpdate:
    """Tests for POST /api/capex-requests/bulk-upload - Updating existing requests"""
    
    def test_bulk_update_existing_request(self, api_client, auth_token):
        """Upload Excel with update sheet updates existing request"""
        # First, create a request to update
        wb_create = Workbook()
        ws_create = wb_create.active
        ws_create.title = "New Requests"
        
        headers = ["Plant *", "Department *", "Requirement Description *", "Quantity", 
                   "Requirement Type *", "Asset Category", "CEA Required", "CEA Type",
                   "Existing CEA Number", "PR Available", "PR Number", "DAP Required", "Justification"]
        for col, header in enumerate(headers, 1):
            ws_create.cell(row=1, column=col, value=header)
        for col in range(1, len(headers) + 1):
            ws_create.cell(row=2, column=col, value="")
        
        data = ["Jaipur", "Quality", "TEST_Update_Target_Machine", 1, "New", "Equipment", "No", "", "", "No", "", "No", "For update test"]
        for col, val in enumerate(data, 1):
            ws_create.cell(row=3, column=col, value=val)
        
        buffer = io.BytesIO()
        wb_create.save(buffer)
        buffer.seek(0)
        
        create_response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload",
            headers={"Authorization": f"Bearer {auth_token}"},
            files={"file": ("create_for_update.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert create_response.status_code == 200, f"Create failed: {create_response.text}"
        create_result = create_response.json()
        assert create_result["summary"]["created"] >= 1, "Should create request for update test"
        
        request_id = create_result["created"][0]["request_id"]
        created_request_ids.append(request_id)
        print(f"Created request for update test: {request_id}")
        
        # Now update it
        wb_update = Workbook()
        ws_update = wb_update.active
        ws_update.title = "Update Existing Requests"
        
        update_headers = ["Request ID *", "CEA Number", "CEA Status", "WBS Number", "PR Number", 
                          "PR Status", "PO Number", "PO Status", "Vendor Name", "Initial Price",
                          "Final Negotiated Price", "Ordered Date", "Expected Delivery Date",
                          "Delivery Status", "Delivery Date", "Installation Date", "Commissioning Date", "Workflow Status"]
        for col, header in enumerate(update_headers, 1):
            ws_update.cell(row=1, column=col, value=header)
        for col in range(1, len(update_headers) + 1):
            ws_update.cell(row=2, column=col, value="")
        
        # Update with PR number and vendor
        update_data = [request_id, "", "", "", "PR-TEST-001", "Approved", "", "", "Test Vendor", 100000, 95000, "", "", "", "", "", "", "PR Approved"]
        for col, val in enumerate(update_data, 1):
            ws_update.cell(row=3, column=col, value=val)
        
        buffer = io.BytesIO()
        wb_update.save(buffer)
        buffer.seek(0)
        
        update_response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload",
            headers={"Authorization": f"Bearer {auth_token}"},
            files={"file": ("update_test.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert update_response.status_code == 200, f"Update failed: {update_response.text}"
        
        update_result = update_response.json()
        assert update_result["summary"]["updated"] >= 1, f"Should update at least 1 request: {update_result}"
        
        # Verify the update by fetching the request
        get_response = api_client.get(
            f"{BASE_URL}/api/capex-requests/{request_id}",
            headers={"Authorization": f"Bearer {auth_token}"}
        )
        assert get_response.status_code == 200
        updated_request = get_response.json()
        assert updated_request.get("pr_number") == "PR-TEST-001", "PR number should be updated"
        assert updated_request.get("workflow_status") == "PR Approved", "Workflow status should be updated"
        print(f"Successfully updated request {request_id}")


class TestBulkUploadValidation:
    """Tests for validation errors in bulk upload"""
    
    def test_missing_required_fields(self, api_client, auth_token):
        """Upload with missing required fields returns row-level errors"""
        wb = Workbook()
        ws = wb.active
        ws.title = "New Requests"
        
        headers = ["Plant *", "Department *", "Requirement Description *", "Quantity", 
                   "Requirement Type *", "Asset Category", "CEA Required", "CEA Type",
                   "Existing CEA Number", "PR Available", "PR Number", "DAP Required", "Justification"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        for col in range(1, len(headers) + 1):
            ws.cell(row=2, column=col, value="")
        
        # Row with missing Plant (required)
        data = ["", "Railway Bearing", "TEST_Missing_Plant", 1, "New", "", "", "", "", "", "", "", ""]
        for col, val in enumerate(data, 1):
            ws.cell(row=3, column=col, value=val)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload",
            headers={"Authorization": f"Bearer {auth_token}"},
            files={"file": ("missing_fields.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        
        result = response.json()
        assert result["summary"]["errors"] >= 1, "Should have at least 1 error"
        assert result["summary"]["created"] == 0, "Should not create any requests"
        
        # Check error message mentions Plant
        errors = result.get("errors", [])
        assert len(errors) > 0, "Should have error details"
        error_text = str(errors[0].get("errors", []))
        assert "Plant" in error_text or "required" in error_text.lower(), f"Error should mention Plant: {error_text}"
        print(f"Validation error correctly returned: {errors}")
    
    def test_invalid_plant_value(self, api_client, auth_token):
        """Upload with invalid plant returns validation error"""
        wb = Workbook()
        ws = wb.active
        ws.title = "New Requests"
        
        headers = ["Plant *", "Department *", "Requirement Description *", "Quantity", 
                   "Requirement Type *", "Asset Category", "CEA Required", "CEA Type",
                   "Existing CEA Number", "PR Available", "PR Number", "DAP Required", "Justification"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        for col in range(1, len(headers) + 1):
            ws.cell(row=2, column=col, value="")
        
        # Row with invalid Plant
        data = ["InvalidPlant", "Railway Bearing", "TEST_Invalid_Plant", 1, "New", "", "", "", "", "", "", "", ""]
        for col, val in enumerate(data, 1):
            ws.cell(row=3, column=col, value=val)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload",
            headers={"Authorization": f"Bearer {auth_token}"},
            files={"file": ("invalid_plant.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        result = response.json()
        assert result["summary"]["errors"] >= 1, "Should have validation error for invalid plant"
        
        errors = result.get("errors", [])
        error_text = str(errors[0].get("errors", []))
        assert "Invalid plant" in error_text or "plant" in error_text.lower(), f"Error should mention invalid plant: {error_text}"
        print(f"Invalid plant error correctly returned: {errors}")
    
    def test_invalid_department_value(self, api_client, auth_token):
        """Upload with invalid department returns validation error"""
        wb = Workbook()
        ws = wb.active
        ws.title = "New Requests"
        
        headers = ["Plant *", "Department *", "Requirement Description *", "Quantity", 
                   "Requirement Type *", "Asset Category", "CEA Required", "CEA Type",
                   "Existing CEA Number", "PR Available", "PR Number", "DAP Required", "Justification"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        for col in range(1, len(headers) + 1):
            ws.cell(row=2, column=col, value="")
        
        # Row with invalid Department
        data = ["Jaipur", "InvalidDepartment", "TEST_Invalid_Dept", 1, "New", "", "", "", "", "", "", "", ""]
        for col, val in enumerate(data, 1):
            ws.cell(row=3, column=col, value=val)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload",
            headers={"Authorization": f"Bearer {auth_token}"},
            files={"file": ("invalid_dept.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        result = response.json()
        assert result["summary"]["errors"] >= 1, "Should have validation error for invalid department"
        print(f"Invalid department error correctly returned: {result.get('errors', [])}")


class TestBulkUploadFileValidation:
    """Tests for file type validation"""
    
    def test_reject_non_excel_file(self, api_client, auth_token):
        """Upload non-Excel file returns error"""
        # Create a text file
        text_content = b"This is not an Excel file"
        
        response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload",
            headers={"Authorization": f"Bearer {auth_token}"},
            files={"file": ("test.txt", io.BytesIO(text_content), "text/plain")}
        )
        
        assert response.status_code == 400, f"Expected 400 for non-Excel file, got {response.status_code}"
        
        result = response.json()
        assert "detail" in result, "Should have error detail"
        assert "xlsx" in result["detail"].lower() or "excel" in result["detail"].lower() or "supported" in result["detail"].lower(), \
            f"Error should mention file type: {result['detail']}"
        print(f"Non-Excel file correctly rejected: {result['detail']}")
    
    def test_reject_pdf_file(self, api_client, auth_token):
        """Upload PDF file returns error"""
        # Create fake PDF header
        pdf_content = b"%PDF-1.4 fake pdf content"
        
        response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload",
            headers={"Authorization": f"Bearer {auth_token}"},
            files={"file": ("test.pdf", io.BytesIO(pdf_content), "application/pdf")}
        )
        
        assert response.status_code == 400, f"Expected 400 for PDF file, got {response.status_code}"
        print("PDF file correctly rejected")


class TestBulkUploadUpdateNotFound:
    """Tests for update with non-existent request ID"""
    
    def test_update_nonexistent_request(self, api_client, auth_token):
        """Update with non-existent request ID returns error"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Update Existing Requests"
        
        update_headers = ["Request ID *", "CEA Number", "CEA Status", "WBS Number", "PR Number", 
                          "PR Status", "PO Number", "PO Status", "Vendor Name", "Initial Price",
                          "Final Negotiated Price", "Ordered Date", "Expected Delivery Date",
                          "Delivery Status", "Delivery Date", "Installation Date", "Commissioning Date", "Workflow Status"]
        for col, header in enumerate(update_headers, 1):
            ws.cell(row=1, column=col, value=header)
        for col in range(1, len(update_headers) + 1):
            ws.cell(row=2, column=col, value="")
        
        # Non-existent request ID
        update_data = ["NONEXISTENT-999", "", "", "", "PR-TEST", "", "", "", "", "", "", "", "", "", "", "", "", ""]
        for col, val in enumerate(update_data, 1):
            ws.cell(row=3, column=col, value=val)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = api_client.post(
            f"{BASE_URL}/api/capex-requests/bulk-upload",
            headers={"Authorization": f"Bearer {auth_token}"},
            files={"file": ("update_nonexistent.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        result = response.json()
        assert result["summary"]["errors"] >= 1, "Should have error for non-existent request"
        assert result["summary"]["updated"] == 0, "Should not update any requests"
        
        errors = result.get("errors", [])
        error_text = str(errors[0].get("errors", []))
        assert "not found" in error_text.lower(), f"Error should mention not found: {error_text}"
        print(f"Non-existent request error correctly returned: {errors}")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
